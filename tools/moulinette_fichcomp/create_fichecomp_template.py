from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import os
import sys
from tkinter import Tk, simpledialog, filedialog


OUTPUT_DEFAULT_NAME = "Fichecomp_template.xlsx"
OUTPUT_DIR = os.path.join("Utilisateur", "MoulinetteExcel")
FIRST_DATA_ROW = 4
LAST_DATA_ROW = 104
HEADER_ROW = 3
INSTRUCTION_ROW = 6
SOURCE_FIRST_ROW = 2

FIXED_VALUES = {
    6: "940140049",
    7: "17",
    9: "940000631",
    11: "ST2",
    12: "06",
    13: '=CONCATENATE(A198,B198,C198,REPT(" ",20-LEN(C198)),D198,REPT(" ",9-LEN(D198)),REPT("0",2-LEN(DAY(E198))),DAY(E198),REPT("0",2-LEN(MONTH(E198))),MONTH(E198),YEAR(E198),F198,G198,REPT(" ",10))',
}

HEADERS = [
    "UF",
    "Libellé - Uf",
    "Colone F du fichcomp rapport 1",
    "Date de naissance",
    "IPP (colone a ajouter)",
    "Finess e-PMSI",
    "Type de prestation",
    "NDA",
    "Numéro FINESS géographique",
    "Date transp. Aller",
    "Code forfait",
    "Classe de distance",
    "Fichcomp",
    "Nombre de kilomètres",
    "Commentaire",
    "Nom fournisseur",
    "Adresse fournisseur ligne 1",
    "Code postal fournisseur",
    "Ville fournisseur",
]


def ask_output_path(default_path):
    root = Tk()
    root.withdraw()
    root.attributes("-topmost", True)
    output_name = simpledialog.askstring(
        "Nom du fichier",
        "Quel nom veux-tu donner au fichier ?",
        initialvalue=OUTPUT_DEFAULT_NAME,
        parent=root,
    )
    if not output_name:
        output_name = OUTPUT_DEFAULT_NAME
    output_name = os.path.basename(output_name)
    if not output_name.lower().endswith(".xlsx"):
        output_name += ".xlsx"

    folder = filedialog.askdirectory(
        title="Choisis le dossier de sauvegarde",
        initialdir=os.path.dirname(default_path),
        parent=root,
    )
    root.destroy()

    if not folder:
        folder = os.path.dirname(default_path)

    return os.path.join(folder, output_name)


def generate_fichecomp_template(output_path):
    wb = Workbook()
    ws = wb.active
    ws.title = "Rendu"

    source = wb.create_sheet("Source")
    source["A1"] = "Date commande"
    source["A1"].font = Font(bold=True)
    source.column_dimensions["A"].width = 18
    for row in range(SOURCE_FIRST_ROW, 102):
        source.cell(row=row, column=1, value="")

    # Title
    ws.merge_cells("A1:R1")
    ws["A1"] = "FICHE COMP - Rendu"
    ws["A1"].font = Font(size=14, bold=True)
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")

    # Header order requested by the user
    # Make the first row a clear import/rendering table
    for idx, h in enumerate(HEADERS, start=1):
        cell = ws.cell(row=HEADER_ROW, column=idx, value=h)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # Keep the hardcoded values in place, as requested.
    # NDA and IPP are intentionally left blank for manual entry.
    for row in range(FIRST_DATA_ROW, LAST_DATA_ROW + 1):
        for col_idx, value in FIXED_VALUES.items():
            ws.cell(row=row, column=col_idx, value=value)
        ws.cell(row=row, column=5, value="")  # IPP to fill manually
        ws.cell(row=row, column=8, value="")  # NDA to fill manually
        source_row = row - (FIRST_DATA_ROW - SOURCE_FIRST_ROW)
        ws.cell(row=row, column=10, value=f"=Source!A{source_row}")
        ws.cell(row=row, column=13, value=FIXED_VALUES[13])

    # Light instruction row
    ws.merge_cells(f"A{INSTRUCTION_ROW}:R{INSTRUCTION_ROW}")
    ws[f"A{INSTRUCTION_ROW}"] = (
        'IPP et NDA sont à saisir à la main. La colonne "Date transp. Aller" est récupérée depuis la feuille Source, colonne "Date commande".'
    )
    ws[f"A{INSTRUCTION_ROW}"].alignment = Alignment(wrap_text=True)
    ws[f"A{INSTRUCTION_ROW}"].font = Font(italic=True)

    # Column widths (make it readable/printable)
    widths = {
        1: 10,
        2: 18,
        3: 24,
        4: 16,
        5: 18,
        6: 14,
        7: 16,
        8: 12,
        9: 24,
        10: 16,
        11: 14,
        12: 16,
        13: 12,
        14: 18,
        15: 18,
        16: 18,
        17: 24,
        18: 18,
        19: 18,
    }
    for col, w in widths.items():
        ws.column_dimensions[get_column_letter(col)].width = w

    # Thin border for table area
    thin = Side(border_style="thin", color="000000")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    for r in range(HEADER_ROW, LAST_DATA_ROW + 1):
        for c in range(1, 20):
            ws.cell(row=r, column=c).border = border

    os.makedirs(os.path.dirname(output_path), exist_ok=True)
    wb.save(output_path)


def main():
    default_path = os.path.join(OUTPUT_DIR, OUTPUT_DEFAULT_NAME)
    out = sys.argv[1] if len(sys.argv) > 1 else ask_output_path(default_path)
    generate_fichecomp_template(out)
    print(f"Fichecomp template created: {out}")


if __name__ == "__main__":
    main()
