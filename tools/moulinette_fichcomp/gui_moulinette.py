from __future__ import annotations

from copy import copy
import os
import threading
import unicodedata
import re
import zipfile
import tempfile
import shutil
from datetime import date, datetime
from pathlib import Path
from queue import Queue, Empty
from typing import Iterable, List, Optional
import tkinter as tk
from tkinter import filedialog, messagebox, ttk
import importlib.util

from openpyxl import Workbook, load_workbook
from openpyxl.cell.cell import MergedCell
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

TARGET_HEADERS = [
    "Date commande",
    "UF",
    "Libellé - Uf",
    "Date paiement - Mnd",
    "Date de naissance",
    "Nom fournisseur",
    "Adresse fournisseur ligne 1",
    "Code postal fournisseur",
    "Ville fournisseur",
    "Nombre de kilomètres",
]

EXCLUDED_PREFIXES = ("~$",)
EXCLUDED_SUFFIXES = ("_clean", "-clean")
EXCEL_EXTENSIONS = {".xlsx", ".xlsm"}
RED_FONT = Font(color="FF0000")
HEADER_KEEP_THRESHOLD = 5
MIN_HEADER_MATCHES = 4
REMOVABLE_ROW_MARKERS = ((4, "nombre :"), (8, "somme :"))
FICHCOMP_HEADERS = [
    "UF",
    "Libellé - Uf",
    "Colone F du fichcomp rapport 1",
    "Date de naissance",
    "IPP(colone a ajouter)",
    "Finess e-PMSI",
    "Type de prestation",
    "NDA",
    "Numéro FINESS géographique",
    "Date transp. Aller",
    "Code forfait",
    "Classe de distance",
    "Fichcomp",
    "Nombre de kilomètres",
    "Commentaire",
    "Nom fournisseur",
    "Adresse fournisseur ligne 1",
    "Code postal fournisseur",
    "Ville fournisseur",
]
FICHCOMP_FIXED_ROW = {
    6: "940140049",
    7: "17",
    9: "940000631",
    10: "Date commande",
    11: "ST2",
    12: "06",
    13: '=CONCATENATE(A198,B198,C198,REPT(" ",20-LEN(C198)),D198,REPT(" ",9-LEN(D198)),REPT("0",2-LEN(DAY(E198))),DAY(E198),REPT("0",2-LEN(MONTH(E198))),MONTH(E198),YEAR(E198),F198,G198,REPT(" ",10))',
}


def normalize_text(value: object) -> str:
    # Rendre le texte comparable en supprimant les accents, les doubles espaces et la casse.
    if value is None:
        return ""
    text = str(value).strip().lower()
    text = unicodedata.normalize("NFKD", text)
    text = "".join(ch for ch in text if not unicodedata.combining(ch))
    text = " ".join(text.split())
    return text


def is_target_header_row(values: Iterable[object]) -> bool:
    # Une ligne d'en-tête complète doit faire apparaître tous les champs attendus.
    normalized_cells = [normalize_text(value) for value in values if normalize_text(value)]
    if not normalized_cells:
        return False
    target_norms = [normalize_text(item) for item in TARGET_HEADERS]
    return all(
        any(target_norm in cell or cell in target_norm for cell in normalized_cells)
        for target_norm in target_norms
    )


def is_partial_header_row(values: Iterable[object]) -> bool:
    # Les lignes partiellement répétées sont acceptées si suffisamment de colonnes clés correspondent.
    normalized_cells = [normalize_text(value) for value in values if normalize_text(value)]
    if not normalized_cells:
        return False

    target_norms = [normalize_text(item) for item in TARGET_HEADERS]
    match_count = 0
    for target_norm in target_norms:
        if any(target_norm in cell or cell in target_norm for cell in normalized_cells):
            match_count += 1

    return match_count >= MIN_HEADER_MATCHES and any(
        target_norm in cell or cell in target_norm
        for cell in normalized_cells
        for target_norm in target_norms[:2]
    )


def output_path_for_source(source: Path, suffix: str = "_clean") -> Path:
    # Le nom de sortie est déterministe pour réécrire la même version clean à chaque traitement.
    return source.with_name(f"{source.stem}{suffix}{source.suffix}")


def save_workbook_with_fallback(workbook: Workbook, output_path: Path) -> Path:
    # Si le fichier cible est déjà ouvert, on essaie un nom libre au lieu d'interrompre le traitement.
    try:
        workbook.save(output_path)
        # Post-traitement: supprimer <calcPr .../> qui peut provoquer des messages de réparation dans Excel.
        try:
            remove_calcPr_from_xlsx(output_path)
        except Exception:
            pass
        return output_path
    except PermissionError:
        fallback_index = 2
        while True:
            fallback_path = output_path.with_name(f"{output_path.stem}_{fallback_index}{output_path.suffix}")
            try:
                workbook.save(fallback_path)
                try:
                    remove_calcPr_from_xlsx(fallback_path)
                except Exception:
                    pass
                return fallback_path
            except PermissionError:
                fallback_index += 1


def remove_calcPr_from_xlsx(xlsx_path: Path) -> None:
    # Retire l'élément calcPr du workbook XML pour éviter des flags non désirés.
    with zipfile.ZipFile(xlsx_path, "r") as zin:
        tmpf = tempfile.NamedTemporaryFile(delete=False)
        tmpf.close()
        with zipfile.ZipFile(tmpf.name, "w", compression=zipfile.ZIP_DEFLATED) as zout:
            for item in zin.infolist():
                data = zin.read(item.filename)
                if item.filename == "xl/workbook.xml":
                    try:
                        text = data.decode("utf-8")
                        new_text = re.sub(r"<calcPr[^>]*/>", "", text)
                        data = new_text.encode("utf-8")
                    except Exception:
                        pass
                zout.writestr(item, data)
    shutil.move(tmpf.name, str(xlsx_path))


def find_header_rows(sheet) -> List[int]:
    # Scanner toutes les lignes pour repérer chaque bloc d'en-tête répété dans la feuille.
    header_rows: List[int] = []
    for row_idx in range(1, sheet.max_row + 1):
        row_values = [
            sheet.cell(row=row_idx, column=col_idx).value for col_idx in range(1, sheet.max_column + 1)
        ]
        if is_target_header_row(row_values) or is_partial_header_row(row_values):
            header_rows.append(row_idx)
    return header_rows


def find_marker_rows(sheet, markers: Iterable[tuple[int, str]]) -> List[int]:
    # Repérer les lignes contenant un libellé de synthèse dans une colonne précise.
    marker_rows: List[int] = []
    for row_idx in range(1, sheet.max_row + 1):
        row_matches_all = True
        for column_idx, marker in markers:
            cell_value = sheet.cell(row=row_idx, column=column_idx).value
            if not cell_value:
                row_matches_all = False
                break
            normalized_cell = normalize_text(cell_value)
            normalized_marker = normalize_text(marker)
            if normalized_cell != normalized_marker:
                row_matches_all = False
                break
        if row_matches_all:
            marker_rows.append(row_idx)
    return marker_rows


def find_marker_rows_in_column(sheet, column_idx: int, marker: str) -> List[int]:
    # Variante simple pour tester un libellé dans une colonne donnée.
    return find_marker_rows(sheet, ((column_idx, marker),))


def copy_sheet_structure(source_sheet, target_sheet) -> None:
    # Reproduire la structure visuelle avant de copier les cellules.
    target_sheet.sheet_format.defaultRowHeight = source_sheet.sheet_format.defaultRowHeight
    target_sheet.sheet_format.defaultColWidth = source_sheet.sheet_format.defaultColWidth
    target_sheet.sheet_view.zoomScale = source_sheet.sheet_view.zoomScale
    target_sheet.freeze_panes = source_sheet.freeze_panes
    if source_sheet.auto_filter.ref:
        target_sheet.auto_filter.ref = source_sheet.auto_filter.ref

    for row_idx, row_dim in source_sheet.row_dimensions.items():
        target_dim = target_sheet.row_dimensions[row_idx]
        target_dim.height = row_dim.height
        target_dim.hidden = row_dim.hidden

    for col_key, col_dim in source_sheet.column_dimensions.items():
        target_dim = target_sheet.column_dimensions[col_key]
        target_dim.width = col_dim.width
        target_dim.hidden = col_dim.hidden


def copy_merged_ranges(source_sheet, target_sheet) -> None:
    # Recopier les cellules fusionnées pour préserver la mise en page du fichier original.
    for merged_range in source_sheet.merged_cells.ranges:
        target_sheet.merge_cells(str(merged_range))


def get_header_rows_to_delete(header_rows: List[int]) -> List[int]:
    # Garder la première occurrence quand l'en-tête est placé tout en haut de la feuille.
    if not header_rows:
        return []
    if header_rows[0] <= HEADER_KEEP_THRESHOLD:
        return header_rows[1:]
    return header_rows


def copy_sheet_content(source_sheet, target_sheet) -> None:
    # Copier valeur, style et annotations cellule par cellule pour conserver l'aspect d'origine.
    copy_sheet_structure(source_sheet, target_sheet)
    for row in source_sheet.iter_rows():
        for source_cell in row:
            if isinstance(source_cell, MergedCell):
                continue
            target_cell = target_sheet.cell(row=source_cell.row, column=source_cell.column)
            target_cell.value = source_cell.value
            if source_cell.has_style:
                target_cell._style = copy(source_cell._style)
            if source_cell.number_format:
                target_cell.number_format = source_cell.number_format
            if source_cell.font:
                target_cell.font = copy(source_cell.font)
            if source_cell.fill:
                target_cell.fill = copy(source_cell.fill)
            if source_cell.border:
                target_cell.border = copy(source_cell.border)
            if source_cell.alignment:
                target_cell.alignment = copy(source_cell.alignment)
            if source_cell.protection:
                target_cell.protection = copy(source_cell.protection)
            if source_cell.hyperlink:
                target_cell._hyperlink = copy(source_cell.hyperlink)
            if source_cell.comment:
                target_cell.comment = copy(source_cell.comment)
    copy_merged_ranges(source_sheet, target_sheet)


def highlight_rows_red(sheet, row_indexes: List[int]) -> None:
    # Signaler visuellement les lignes supprimées dans la version originale.
    for row_idx in row_indexes:
        for cell in sheet[row_idx]:
            cell.font = copy(cell.font)
            cell.font = cell.font.copy(color=RED_FONT.color.rgb)


def clean_modified_sheet(sheet) -> int:
    # Recalculer les dates d'abord, puis supprimer les lignes de total, puis les en-têtes.
    propagate_dates_in_column_b(sheet)

    removable_rows = find_marker_rows(sheet, REMOVABLE_ROW_MARKERS)
    deleted_rows = 0
    for row_idx in reversed(removable_rows):
        sheet.delete_rows(row_idx, 1)
        deleted_rows += 1

    header_rows = find_header_rows(sheet)
    header_rows_to_delete = get_header_rows_to_delete(header_rows)
    for row_idx in reversed(header_rows_to_delete):
        sheet.delete_rows(row_idx, 1)
        deleted_rows += 1

    return deleted_rows


def is_date_like(value: object) -> bool:
    # Détecter les valeurs qui représentent une date, quelle que soit leur forme d'origine.
    if value is None:
        return False
    if isinstance(value, (datetime, date)):
        return True
    if isinstance(value, str):
        text = value.strip()
        if not text:
            return False
        for fmt in (
            "%d/%m/%Y",
            "%d/%m/%Y %H:%M:%S",
            "%Y-%m-%d",
            "%Y-%m-%d %H:%M:%S",
            "%d-%m-%Y",
            "%d.%m.%Y",
        ):
            try:
                datetime.strptime(text, fmt)
                return True
            except ValueError:
                continue
    return False


def propagate_dates_in_column_b(sheet) -> int:
    # Convertir les dates trouvées en vraies dates Excel puis remplir les cellules vides du dessous.
    def parse_date_value(val: object) -> Optional[datetime]:
        if val is None:
            return None
        if isinstance(val, datetime):
            return val
        if isinstance(val, date):
            return datetime(val.year, val.month, val.day)
        if isinstance(val, str):
            text = val.strip()
            if not text:
                return None
            for fmt in (
                "%d/%m/%Y",
                "%d/%m/%Y %H:%M:%S",
                "%Y-%m-%d",
                "%Y-%m-%d %H:%M:%S",
                "%d-%m-%Y",
                "%d.%m.%Y",
            ):
                try:
                    return datetime.strptime(text, fmt)
                except ValueError:
                    continue
        return None

    last_date: Optional[datetime] = None
    filled_count = 0
    # traverse rows and propagate parsed datetime objects into column B
    for row_idx in range(1, sheet.max_row + 1):
        cell = sheet.cell(row=row_idx, column=2)
        if isinstance(cell, MergedCell):
            continue

        value = cell.value
        parsed = parse_date_value(value)
        if parsed is not None:
            last_date = parsed
            # ensure the cell is a proper date type and formatted
            cell.value = parsed
            try:
                cell.number_format = "DD/MM/YYYY"
            except Exception:
                pass
            continue

        if last_date is not None and (value is None or (isinstance(value, str) and not value.strip())):
            cell.value = last_date
            try:
                cell.number_format = "DD/MM/YYYY"
            except Exception:
                pass
            filled_count += 1

    # Ajuster la largeur de la colonne B pour éviter l'affichage ######## dans Excel.
    try:
        col = "B"
        max_len = 0
        for row_idx in range(1, sheet.max_row + 1):
            cell = sheet.cell(row=row_idx, column=2)
            if cell.value is None:
                continue
            # format value as displayed string
            try:
                if isinstance(cell.value, (datetime, date)):
                    text = cell.value.strftime("%d/%m/%Y")
                else:
                    text = str(cell.value)
            except Exception:
                text = str(cell.value)
            max_len = max(max_len, len(text))
        # add small padding
        width = max(10, min(40, max_len + 2))
        sheet.column_dimensions[col].width = width
    except Exception:
        pass

    return filled_count


def process_workbook(source_path: Path, progress_callback=None) -> Path:
    # Construire un classeur de sortie avec une feuille originale et une feuille nettoyée par onglet.
    keep_vba = source_path.suffix.lower() == ".xlsm"
    workbook = load_workbook(source_path, keep_vba=keep_vba)

    output_workbook = Workbook()
    output_workbook.remove(output_workbook.active)

    total_deleted_rows = 0
    for source_sheet in workbook.worksheets:
        header_rows = find_header_rows(source_sheet)
        rows_to_delete = get_header_rows_to_delete(header_rows)
        rows_to_delete.extend(find_marker_rows(source_sheet, REMOVABLE_ROW_MARKERS))
        rows_to_delete = sorted(set(rows_to_delete))

        original_title = make_sheet_title(output_workbook, f"{source_sheet.title} original")
        modified_title = make_sheet_title(output_workbook, f"{source_sheet.title} modifie")

        original_sheet = output_workbook.create_sheet(title=original_title)
        modified_sheet = output_workbook.create_sheet(title=modified_title)

        copy_sheet_content(source_sheet, original_sheet)
        copy_sheet_content(source_sheet, modified_sheet)

        highlight_rows_red(original_sheet, rows_to_delete)
        total_deleted_rows += clean_modified_sheet(modified_sheet)

    add_fichcomp_sheet(output_workbook)

    modified_sheet_index = next(
        (index for index, sheet in enumerate(output_workbook.worksheets) if sheet.title.endswith(" modifie")),
        0,
    )
    output_workbook.active = modified_sheet_index

    # Certaines versions d'Excel signalent des réparations si le flag
    # fullCalcOnLoad est présent/mal valorisé dans le workbook XML.
    # Le désactiver ici tout en conservant les formules réduit les risques
    # de message de réparation côté Excel.
    try:
        output_workbook.calc_properties.fullCalcOnLoad = False
    except Exception:
        pass

    output_path = output_path_for_source(source_path)
    output_path = save_workbook_with_fallback(output_workbook, output_path)

    if progress_callback:
        progress_callback(
            f"{source_path.name} -> {output_path.name} | lignes supprimées dans la version modifiée: {total_deleted_rows}"
        )
    return output_path


def add_fichcomp_sheet(workbook: Workbook) -> None:
    # Ajouter la page Fichcomp demandée juste après le nettoyage.
    modified_sheet = next(
        (sheet for sheet in workbook.worksheets if sheet.title.endswith(" modifie")),
        None,
    )
    sheet = workbook.create_sheet(title=make_sheet_title(workbook, "Fichcomp"))
    sheet.sheet_view.showGridLines = False

    header_fill = PatternFill("solid", fgColor="B8CCE4")
    orange_fill = PatternFill("solid", fgColor="F4B183")
    yellow_fill = PatternFill("solid", fgColor="FFF200")
    border = Border(
        left=Side(style="thin", color="808080"),
        right=Side(style="thin", color="808080"),
        top=Side(style="thin", color="808080"),
        bottom=Side(style="thin", color="808080"),
    )

    for col_idx, header in enumerate(FICHCOMP_HEADERS, start=1):
        cell = sheet.cell(row=1, column=col_idx, value=header)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.font = Font(size=10)
        cell.border = border
        if col_idx in {5, 15}:
            cell.fill = orange_fill
        elif col_idx == 13:
            cell.fill = yellow_fill
        else:
            cell.fill = header_fill
        sheet.column_dimensions[cell.column_letter].width = max(12, min(26, len(header) + 2))

    if modified_sheet is not None:
        source_rows = range(4, modified_sheet.max_row + 1)
    else:
        source_rows = range(4, 5)

    for output_row, source_row in enumerate(source_rows, start=2):
        row_values = {
            1: f"='{modified_sheet.title}'!C{source_row}" if modified_sheet else None,
            2: f"='{modified_sheet.title}'!D{source_row}" if modified_sheet else None,
            3: f"='{modified_sheet.title}'!F{source_row}" if modified_sheet else None,
            4: f"='{modified_sheet.title}'!G{source_row}" if modified_sheet else None,
            5: None,
            6: FICHCOMP_FIXED_ROW[6],
            7: FICHCOMP_FIXED_ROW[7],
            8: None,
            9: FICHCOMP_FIXED_ROW[9],
            10: f"='{modified_sheet.title}'!B{source_row}" if modified_sheet else None,
            11: FICHCOMP_FIXED_ROW[11],
            12: FICHCOMP_FIXED_ROW[12],
            13: f'=CONCATENATE(A{198 + (output_row - 2)},B{198 + (output_row - 2)},C{198 + (output_row - 2)},REPT(" ",20-LEN(C{198 + (output_row - 2)})),D{198 + (output_row - 2)},REPT(" ",9-LEN(D{198 + (output_row - 2)})),REPT("0",2-LEN(DAY(E{198 + (output_row - 2)}))),DAY(E{198 + (output_row - 2)}),REPT("0",2-LEN(MONTH(E{198 + (output_row - 2)}))),MONTH(E{198 + (output_row - 2)}),YEAR(E{198 + (output_row - 2)}),F{198 + (output_row - 2)},G{198 + (output_row - 2)},REPT(" ",10))',
            14: f"='{modified_sheet.title}'!L{source_row}" if modified_sheet else None,
            15: None,
            16: f"='{modified_sheet.title}'!H{source_row}" if modified_sheet else None,
            17: f"='{modified_sheet.title}'!I{source_row}" if modified_sheet else None,
            18: f"='{modified_sheet.title}'!J{source_row}" if modified_sheet else None,
            19: f"='{modified_sheet.title}'!K{source_row}" if modified_sheet else None,
        }

        for col_idx in range(1, len(FICHCOMP_HEADERS) + 1):
            cell = sheet.cell(row=output_row, column=col_idx, value=row_values.get(col_idx))
            cell.border = border
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            if col_idx == 13:
                cell.fill = yellow_fill
            elif col_idx in {5, 15}:
                cell.fill = orange_fill
            else:
                cell.fill = PatternFill("solid", fgColor="FFFFFF")

    if modified_sheet is None:
        sample_row = 2
        for col_idx in range(1, len(FICHCOMP_HEADERS) + 1):
            cell = sheet.cell(row=sample_row, column=col_idx, value=None)
            cell.border = border
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            if col_idx == 13:
                cell.fill = yellow_fill
            elif col_idx in {5, 15}:
                cell.fill = orange_fill
            else:
                cell.fill = PatternFill("solid", fgColor="FFFFFF")

    sheet.freeze_panes = "A2"
    sheet.row_dimensions[1].height = 40
    sheet.row_dimensions[2].height = 24


def make_sheet_title(workbook: Workbook, title: str) -> str:
    # Garantir un titre de feuille valide et unique dans la limite imposée par Excel.
    base = title[:31]
    candidate = base
    counter = 2
    existing = {sheet.title for sheet in workbook.worksheets}
    while candidate in existing:
        suffix = f"_{counter}"
        candidate = f"{base[: 31 - len(suffix)]}{suffix}"
        counter += 1
    return candidate


def find_excel_files(folder: Path) -> List[Path]:
    # Ignorer les temporaires, les copies clean et les fichiers hors Excel.
    files = []
    for path in sorted(folder.rglob("*")):
        if not path.is_file():
            continue
        if path.name.startswith(EXCLUDED_PREFIXES):
            continue
        if path.stem.endswith(EXCLUDED_SUFFIXES):
            continue
        if path.suffix.lower() not in EXCEL_EXTENSIONS:
            continue
        files.append(path)
    return files


class CleanerApp:
    def __init__(self, root: tk.Tk) -> None:
        # Construire l'interface principale et préparer la file de messages du worker.
        self.root = root
        self.root.title("Moulinette Excel")
        self.root.geometry("720x420")
        self.root.minsize(640, 380)

        self.queue: Queue[str | tuple[str, str]] = Queue()
        self.worker_running = False
        self.selected_files: List[Path] = []
        self.selected_folder: Optional[Path] = None

        notebook = ttk.Notebook(root)
        notebook.pack(fill="both", expand=True)

        main = ttk.Frame(notebook, padding=16)
        fichcomp = ttk.Frame(notebook, padding=12)
        notebook.add(main, text="Nettoyage")
        notebook.add(fichcomp, text="Fichcomp")

        title = ttk.Label(main, text="Nettoyage des fichiers Excel", font=("Segoe UI", 16, "bold"))
        title.pack(anchor="w")

        self.folder_var = tk.StringVar(value=f"Dossier détecté : {self.project_folder()}")
        folder_label = ttk.Label(main, textvariable=self.folder_var)
        folder_label.pack(anchor="w", pady=(8, 16))

        pick_frame = ttk.Frame(main)
        pick_frame.pack(anchor="w", pady=(0, 8))

        ttk.Button(pick_frame, text="Choisir fichier(s)", command=self.pick_files).pack(side="left")
        ttk.Button(pick_frame, text="Choisir un dossier", command=self.pick_folder).pack(
            side="left", padx=(8, 0)
        )
        ttk.Button(pick_frame, text="Mode automatique", command=self.clear_selection).pack(
            side="left", padx=(8, 0)
        )

        self.selection_var = tk.StringVar(value="Mode actuel : dossier Fichier (automatique)")
        ttk.Label(main, textvariable=self.selection_var).pack(anchor="w", pady=(0, 8))

        self.run_button = ttk.Button(main, text="Lancer le nettoyage", command=self.start)
        self.run_button.pack(anchor="w")

        self.progress = ttk.Progressbar(main, mode="indeterminate")
        self.progress.pack(fill="x", pady=(12, 12))

        log_frame = ttk.LabelFrame(main, text="Journal")
        log_frame.pack(fill="both", expand=True)

        self.log = tk.Text(log_frame, height=12, wrap="word", state="disabled")
        self.log.pack(fill="both", expand=True, padx=8, pady=8)

        hint = ttk.Label(
            main,
            text="Le programme crée une copie *_clean.xlsx avec 2 feuilles : version modifiée et original en rouge.",
        )
        hint.pack(anchor="w", pady=(10, 0))

        self.build_fichcomp_page(fichcomp)

        self.root.after(200, self.poll_queue)

    def build_fichcomp_page(self, parent: ttk.Frame) -> None:
        # Afficher la page demandée au lancement, avec les champs et valeurs fixes du gabarit.
        intro = ttk.Label(
            parent,
            text="Fichecomp - page de saisie",
            font=("Segoe UI", 14, "bold"),
        )
        intro.pack(anchor="w", pady=(0, 8))

        canvas = tk.Canvas(parent, highlightthickness=0)
        scrollbar = ttk.Scrollbar(parent, orient="horizontal", command=canvas.xview)
        canvas.configure(xscrollcommand=scrollbar.set)
        scrollbar.pack(side="bottom", fill="x")
        canvas.pack(side="top", fill="both", expand=True)

        body = ttk.Frame(canvas)
        canvas.create_window((0, 0), window=body, anchor="nw")

        self.fichcomp_entries: dict[str, tk.Entry] = {}

        for col, header in enumerate(FICHCOMP_HEADERS):
            lbl = ttk.Label(body, text=header, anchor="center", justify="center", wraplength=110)
            lbl.grid(row=0, column=col, sticky="nsew", padx=1, pady=1)
            body.columnconfigure(col, weight=1)

        sample = ["" for _ in FICHCOMP_HEADERS]
        sample[5] = FICHCOMP_FIXED_ROW[6]
        sample[6] = FICHCOMP_FIXED_ROW[7]
        sample[8] = FICHCOMP_FIXED_ROW[9]
        sample[9] = FICHCOMP_FIXED_ROW[10]
        sample[10] = FICHCOMP_FIXED_ROW[11]
        sample[11] = FICHCOMP_FIXED_ROW[12]
        sample[12] = FICHCOMP_FIXED_ROW[13]

        for col, value in enumerate(sample):
            entry = ttk.Entry(body, width=max(10, len(str(value)) + 2))
            if value:
                entry.insert(0, value)
            entry.grid(row=1, column=col, sticky="nsew", padx=1, pady=(0, 8))
            self.fichcomp_entries[FICHCOMP_HEADERS[col]] = entry

        self.fichcomp_entries["IPP(colone a ajouter)"].insert(0, "")
        self.fichcomp_entries["NDA"].insert(0, "")

        self.fichcomp_entries["Date transp. Aller"].delete(0, "end")
        self.fichcomp_entries["Date transp. Aller"].insert(0, "Date commande")

        footer = ttk.Label(
            parent,
            text="IPP et NDA restent à saisir à la main. Date transp. Aller est alimentée depuis la colonne Date commande.",
        )
        footer.pack(anchor="w", pady=(8, 0))

        ttk.Button(parent, text="Créer le fichier Fichecomp", command=self.create_fichcomp_workbook).pack(
            anchor="w", pady=(10, 0)
        )

        body.bind("<Configure>", lambda event: canvas.configure(scrollregion=canvas.bbox("all")))

    def create_fichcomp_workbook(self) -> None:
        # Générer le classeur fichecomp à partir de la page affichée.
        try:
            generator_path = Path(__file__).resolve().parents[2] / "code" / "create_fichecomp_template.py"
            spec = importlib.util.spec_from_file_location("create_fichecomp_template", generator_path)
            if spec is None or spec.loader is None:
                raise RuntimeError("spec introuvable")
            module = importlib.util.module_from_spec(spec)
            spec.loader.exec_module(module)
        except Exception:
            messagebox.showerror("Moulinette Excel", "Impossible de charger le générateur Fichecomp.")
            return

        output_path = filedialog.asksaveasfilename(
            title="Enregistrer le fichier Fichecomp",
            defaultextension=".xlsx",
            filetypes=[("Excel", "*.xlsx")],
            initialfile="Fichecomp_template.xlsx",
        )
        if not output_path:
            return
        module.generate_fichecomp_template(Path(output_path))
        messagebox.showinfo("Moulinette Excel", "Fichier Fichecomp créé.")

    def project_folder(self) -> Path:
        # Résoudre le dossier Fichier quel que soit le mode de lancement de l'application.
        import sys

        if getattr(sys, "frozen", False):
            base = Path(sys.executable).resolve().parent
        else:
            file_dir = Path(__file__).resolve().parent
            base = file_dir.parent if file_dir.name.lower() == "code" else file_dir

        folder = base / "Fichier"
        folder.mkdir(parents=True, exist_ok=True)
        return folder

    def append_log(self, message: str) -> None:
        # Ajouter une ligne dans le journal visible par l'utilisateur.
        self.log.configure(state="normal")
        self.log.insert("end", message + "\n")
        self.log.see("end")
        self.log.configure(state="disabled")

    def pick_files(self) -> None:
        # Laisser l'utilisateur choisir un ensemble précis de fichiers Excel.
        paths = filedialog.askopenfilenames(
            title="Choisir un ou plusieurs fichiers Excel",
            filetypes=[("Excel", "*.xlsx *.xlsm")],
        )
        if not paths:
            return
        self.selected_files = [Path(item) for item in paths]
        self.selected_folder = None
        self.selection_var.set(f"Mode actuel : {len(self.selected_files)} fichier(s) sélectionné(s)")

    def pick_folder(self) -> None:
        # Laisser l'utilisateur traiter tout un dossier à la place du mode automatique.
        selected = filedialog.askdirectory(title="Choisir un dossier contenant des fichiers Excel")
        if not selected:
            return
        self.selected_folder = Path(selected)
        self.selected_files = []
        self.selection_var.set(f"Mode actuel : dossier sélectionné -> {self.selected_folder}")

    def clear_selection(self) -> None:
        # Revenir au comportement par défaut basé sur le dossier Fichier.
        self.selected_files = []
        self.selected_folder = None
        self.selection_var.set("Mode actuel : dossier Fichier (automatique)")

    def poll_queue(self) -> None:
        # Récupérer les messages du worker sur le thread UI pour mettre à jour l'écran proprement.
        try:
            while True:
                item = self.queue.get_nowait()
                kind, value = item
                if kind == "log":
                    self.append_log(value)
                elif kind == "done":
                    message, folder = value
                    self.append_log(message)
                    self.progress.stop()
                    self.run_button.configure(state="normal")
                    self.worker_running = False
                    messagebox.showinfo("Moulinette Excel", "Nettoyage terminé.")
                    if folder is not None:
                        self.open_folder(folder)
        except Empty:
            pass
        finally:
            self.root.after(200, self.poll_queue)

    def get_target_files(self) -> List[Path]:
        # Priorité à la sélection explicite, puis au dossier choisi, puis au dossier automatique.
        if self.selected_files:
            return [
                path
                for path in self.selected_files
                if path.exists() and path.suffix.lower() in EXCEL_EXTENSIONS
            ]
        if self.selected_folder:
            return find_excel_files(self.selected_folder)
        return find_excel_files(self.project_folder())

    def start(self) -> None:
        # Bloquer les relances concurrentes et démarrer le travail en arrière-plan.
        if self.worker_running:
            return

        excel_files = self.get_target_files()
        if not excel_files:
            messagebox.showinfo("Moulinette Excel", "Aucun fichier Excel trouvé dans la sélection courante.")
            return

        self.worker_running = True
        self.run_button.configure(state="disabled")
        self.progress.start(12)
        self.append_log(f"Démarrage du traitement pour {len(excel_files)} fichier(s).")

        thread = threading.Thread(target=self.run_worker, args=(excel_files,), daemon=True)
        thread.start()

    def run_worker(self, files: List[Path]) -> None:
        # Exécuter le traitement sans bloquer l'interface et mémoriser un dossier à rouvrir.
        if self.selected_files:
            location = "sélection de fichiers"
        elif self.selected_folder:
            location = str(self.selected_folder)
        else:
            location = str(self.project_folder())

        opened_folder: Optional[Path] = None

        for path in files:
            try:
                self.queue.put(("log", f"Traitement de {path.name}..."))
                process_workbook(path, progress_callback=lambda msg: self.queue.put(("log", msg)))
                opened_folder = path.parent
            except Exception as exc:  # pragma: no cover - surfaced in UI
                self.queue.put(("log", f"Erreur sur {path.name} : {exc}"))
        self.queue.put(("done", (f"Terminé. Fichiers traités depuis : {location}", opened_folder)))

    def open_folder(self, folder: Path) -> None:
        # Ouvrir le dossier traité avec l'explorateur Windows.
        try:
            os.startfile(folder)
        except OSError:
            self.append_log(f"Impossible d'ouvrir le dossier : {folder}")


def main() -> None:
    root = tk.Tk()
    try:
        ttk.Style().theme_use("clam")
    except tk.TclError:
        pass
    CleanerApp(root)
    root.mainloop()


if __name__ == "__main__":
    main()
