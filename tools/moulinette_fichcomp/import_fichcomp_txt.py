import sys
import os
from openpyxl import Workbook
from datetime import datetime


def parse_txt_to_rows(path):
    rows = []
    with open(path, "r", encoding="utf-8", errors="ignore") as f:
        for line in f:
            s = line.rstrip("\n")
            if not s:
                continue
            parts = s.split()
            if len(parts) >= 2:
                rows.append((parts[0], parts[1], " ".join(parts[2:]) if len(parts) > 2 else ""))
            else:
                rows.append((parts[0], "", ""))
    return rows


def create_example_fiche(rows, out_path):
    wb = Workbook()
    # Raw sheet
    raw = wb.active
    raw.title = "Raw"
    for i, r in enumerate(rows, start=1):
        raw.cell(row=i, column=1, value=r[0])
        raw.cell(row=i, column=2, value=r[1])
        if r[2]:
            raw.cell(row=i, column=3, value=r[2])

    # FicheComp formatted sheet
    fiche = wb.create_sheet("FicheComp")
    fiche.cell(row=1, column=1, value="FICHE COMP - Exemple")
    headers = ["N°", "Désignation", "Unité", "Nombre", "PU HT", "Total HT", "TVA", "Somme"]
    for c, h in enumerate(headers, start=1):
        fiche.cell(row=3, column=c, value=h)

    start = 4
    for i, r in enumerate(rows, start=start):
        fiche.cell(row=i, column=1, value=i - start + 1)
        # Place token1 into Désignation and token2 into Somme-code column for example
        fiche.cell(row=i, column=2, value=r[0])
        fiche.cell(row=i, column=8, value=r[1])

    summary_row = start + len(rows) + 2
    fiche.cell(row=summary_row, column=4, value="Nombre :")
    fiche.cell(row=summary_row, column=8, value="Somme :")

    os.makedirs(os.path.dirname(out_path), exist_ok=True)
    wb.save(out_path)


def main():
    if len(sys.argv) < 2:
        print("Usage: import_fichcomp_txt.py <path-to-txt> [out_xlsx]")
        sys.exit(1)
    src = sys.argv[1]
    out = sys.argv[2] if len(sys.argv) > 2 else None
    if not out:
        base = os.path.splitext(os.path.basename(src))[0]
        ts = datetime.now().strftime("%Y%m%d_%H%M%S")
        out = os.path.join("Utilisateur", "MoulinetteExcel", f"Example_Fichecomp_{base}_{ts}.xlsx")

    rows = parse_txt_to_rows(src)
    create_example_fiche(rows, out)
    print("Example fichecomp created:", out)


if __name__ == "__main__":
    main()
