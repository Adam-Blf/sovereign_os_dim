"""
Exporter Excel -> FICHCOMP / FICHDMI fixed-width text

Usage:
  py -3 export_to_fichcomp.py input.xlsx output.txt --type med --finess 000000001

The script reads the first sheet and looks for headers (case-insensitive):
  finess, admin_stay, code, nombre, date
If headers are not found, it will read the first five columns as the above fields.

Field formatting (FICHCOMP medication):
  FINESS: 9 chars, zero-padded left
  N° administratif de séjour: 20 chars, space-padded right
  Code UCD: 9 chars, zero-padded left
  Nombre d'UCD administrées: 7 chars, integer = value * 1000, zero-padded left (4+3)
  Date: 8 chars ddmmyyyy or 8 spaces if empty

For DMI (--type dmi) the layout is similar but Nombre width differs (4 chars) and Date position differs.
"""

from openpyxl import load_workbook
import os
import argparse


def fmt_str(value, length, zpad=False):
    s = "" if value is None else str(value)
    s = s.strip()
    if zpad:
        s = s.zfill(length)
    else:
        if len(s) > length:
            s = s[:length]
        s = s.ljust(length)
    return s


def fmt_nombre(value, width, decimals=3):
    if value is None or str(value).strip() == "":
        return "0".zfill(width)
    try:
        v = float(value)
    except Exception:
        # try to clean comma
        try:
            v = float(str(value).replace(",", "."))
        except Exception:
            v = 0.0
    scaled = int(round(v * (10**decimals)))
    return str(scaled).zfill(width)


def fmt_date(value):
    if value is None:
        return " " * 8
    s = str(value).strip()
    if not s:
        return " " * 8
    # Accept dd/mm/YYYY or ddmmyyyy or Excel date objects
    s = s.replace("/", "").replace("-", "")
    if len(s) == 8 and s.isdigit():
        return s
    # fallback: try to parse
    from datetime import datetime

    for fmt in ("%d%m%Y", "%d%m%y", "%Y%m%d", "%d%m%Y"):
        try:
            dt = datetime.strptime(s, fmt)
            return dt.strftime("%d%m%Y")
        except Exception:
            pass
    return " " * 8


def export_med(sheet, out_f, default_finess="000000001"):
    # find headers
    rows = list(sheet.iter_rows(values_only=True))
    if not rows:
        return 0
    # Try to find header row in the first 20 rows by looking for known header keywords
    header_keywords = (
        "finess",
        "n° finess",
        "admin_stay",
        "n° administratif de séjour",
        "administratif",
        "code",
        "ucd",
        "lpp",
        "nombre",
        "date",
        "désignation",
        "n°",
        "nombre",
    )
    header_idx = None
    for idx in range(min(20, len(rows))):
        rowvals = [str(c).strip().lower() if c is not None else "" for c in rows[idx]]
        matches = sum(1 for h in header_keywords if h in " ".join(rowvals))
        if matches >= 2:
            header_idx = idx
            break
    if header_idx is not None:
        headers = [str(c).strip().lower() if c is not None else "" for c in rows[header_idx]]
        start = header_idx + 1
    else:
        headers = [str(c).strip().lower() if c is not None else "" for c in rows[0]]
        start = 0
    count = 0
    for r in rows[start:]:
        # mapping
        def get_by_name(*names):
            for h in names:
                if h in headers:
                    try:
                        return r[headers.index(h)]
                    except Exception:
                        return ""
            return ""

        # Prefer named columns when headers were detected; otherwise fallback to positional
        finess = get_by_name("finess", "n° finess") or default_finess
        admin = get_by_name("admin_stay", "n° administratif de séjour", "administratif") or (
            r[1] if len(r) > 1 else ""
        )
        code = get_by_name("code", "ucd", "lpp") or (r[2] if len(r) > 2 else "")
        nombre = get_by_name("nombre") or (r[3] if len(r) > 3 else "")
        date = get_by_name("date") or (r[4] if len(r) > 4 else "")

        line = ""
        line += fmt_str(str(finess), 9, zpad=True)
        line += fmt_str(admin, 20, zpad=False)
        line += fmt_str(str(code), 9, zpad=True)
        line += fmt_nombre(nombre, 7, decimals=3)
        line += fmt_date(date)
        out_f.write(line + "\n")
        count += 1
    return count


def export_dmi(sheet, out_f, default_finess="000000001"):
    # Similar to med but Nombre width=4 and date width=8 at columns 36-43 per spec
    rows = list(sheet.iter_rows(values_only=True))
    if not rows:
        return 0
    headers = [str(c).strip().lower() if c is not None else "" for c in rows[0]]
    has_headers = any(
        h in ("finess", "admin_stay", "administratif", "code", "lpp", "nombre", "date") for h in headers
    )
    start = 1 if has_headers else 0
    count = 0
    for r in rows[start:]:
        if has_headers:

            def get(h):
                try:
                    idx = headers.index(h)
                    return r[idx]
                except ValueError:
                    return ""

            finess = get("finess") or default_finess
            admin = get("admin_stay") or ""
            code = get("code") or get("lpp") or ""
            nombre = get("nombre")
            date = get("date")
        else:
            finess = r[0] or default_finess
            admin = r[1] if len(r) > 1 else ""
            code = r[2] if len(r) > 2 else ""
            nombre = r[3] if len(r) > 3 else ""
            date = r[4] if len(r) > 4 else ""

        line = ""
        line += fmt_str(str(finess), 9, zpad=True)
        line += fmt_str(admin, 20, zpad=False)
        line += fmt_str(str(code), 9, zpad=True)
        # Nombre posées: 4 chars (integer)
        try:
            n = int(float(nombre)) if nombre not in (None, "") else 0
        except Exception:
            n = 0
        line += str(n).zfill(4)
        line += fmt_date(date)
        out_f.write(line + "\n")
        count += 1
    return count


def main():
    p = argparse.ArgumentParser(description="Export Excel to FICHCOMP/FICHDMI fixed-width text")
    p.add_argument("input", help="input xlsx")
    p.add_argument("output", help="output txt")
    p.add_argument("--type", choices=("med", "dmi"), default="med")
    p.add_argument("--finess", default="000000001")
    args = p.parse_args()

    wb = load_workbook(args.input, data_only=True)
    sheet = wb.active
    os.makedirs(os.path.dirname(args.output), exist_ok=True)
    with open(args.output, "w", encoding="utf-8") as out_f:
        if args.type == "med":
            n = export_med(sheet, out_f, default_finess=args.finess)
        else:
            n = export_dmi(sheet, out_f, default_finess=args.finess)
    print(f"Wrote {n} records to {args.output}")


if __name__ == "__main__":
    main()
