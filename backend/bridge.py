# ══════════════════════════════════════════════════════════════════════════════
#  SOVEREIGN OS DIM — HTTP REST Bridge
# ══════════════════════════════════════════════════════════════════════════════
#  Author  : Adam Beloucif
#  Project : Sovereign OS V32.0 — Station DIM GHT Sud Paris
#
#  Description:
#    Pont HTTP REST exposant le moteur ATIH (DataProcessor) à des
#    applications tierces — en particulier une application PHP déployée
#    sur le même poste ou le même LAN hospitalier.
#
#    L'API pywebview (backend/api.py) reste réservée au frontend embarqué.
#    Ce bridge ré-expose les mêmes capacités métier en HTTP JSON, avec :
#      - Authentification par jeton Bearer (SOVEREIGN_BRIDGE_TOKEN)
#      - Binding par défaut sur 127.0.0.1 (ne sort jamais du poste)
#      - CORS restrictif (liste blanche d'origines)
#
#  Endpoints principaux :
#    GET  /health                      — heartbeat (pas d'auth)
#    GET  /api/matrix                  — 23 formats ATIH
#    POST /api/identify                — identifie un fichier par son nom
#    POST /api/scan                    — scanne des dossiers
#    POST /api/process                 — scan + process en un appel
#    GET  /api/collisions              — liste les collisions MPI
#    POST /api/resolve                 — set_pivot manuel OU auto-résolution
#    GET  /api/stats                   — stats dashboard
#    POST /api/export                  — export CSV vers un dossier
#    POST /api/export-sanitized        — export .txt purifié
#    POST /api/inspect                 — inspection ligne par ligne
#    POST /api/import-csv              — lecture d'un CSV externe
#    POST /api/import-excel            — lecture d'un classeur Excel (.xlsx)
#    POST /api/chart-from-excel        — série agrégée prête pour Chart.js
#    POST /api/reset                   — remise à zéro
#
#  Usage:
#    python -m backend.bridge --host 127.0.0.1 --port 8765
#    SOVEREIGN_BRIDGE_TOKEN=secret python -m backend.bridge
# ══════════════════════════════════════════════════════════════════════════════

from __future__ import annotations

import argparse
import os
import secrets
from functools import wraps
from typing import Callable

from flask import Flask, jsonify, request

from backend.data_processor import DataProcessor


# ──────────────────────────────────────────────────────────────────────────────
# CONFIGURATION — Variables d'environnement
# ──────────────────────────────────────────────────────────────────────────────

DEFAULT_HOST = os.environ.get("SOVEREIGN_BRIDGE_HOST", "127.0.0.1")
DEFAULT_PORT = int(os.environ.get("SOVEREIGN_BRIDGE_PORT", "8765"))
# Jeton partagé entre le bridge et le client PHP. Un jeton vide DÉSACTIVE
# l'authentification — à n'utiliser qu'en développement local.
BRIDGE_TOKEN = os.environ.get("SOVEREIGN_BRIDGE_TOKEN", "")
# Origines autorisées pour CORS (séparées par des virgules). "*" = tout
# autoriser (déconseillé en prod). Par défaut : localhost uniquement.
ALLOWED_ORIGINS = [
    o.strip()
    for o in os.environ.get(
        "SOVEREIGN_BRIDGE_ORIGINS",
        "http://127.0.0.1,http://localhost",
    ).split(",")
    if o.strip()
]


# ──────────────────────────────────────────────────────────────────────────────
# INSTANCE UNIQUE DU PROCESSEUR — partagé entre toutes les requêtes
# ──────────────────────────────────────────────────────────────────────────────
# Pourquoi un singleton ? Le MPI (Master Patient Index) est un état construit
# au fil des scans. Chaque requête HTTP doit voir l'état accumulé, sinon
# l'API PHP ne pourrait jamais résoudre une collision trouvée au scan précédent.

_processor = DataProcessor()
_current_folders: list[str] = []
_current_files: list[dict] = []


def _reset_state() -> None:
    """Remet le bridge dans un état vierge (nouveau processor)."""
    global _processor, _current_folders, _current_files
    _processor = DataProcessor()
    _current_folders = []
    _current_files = []


# ──────────────────────────────────────────────────────────────────────────────
# FLASK APP
# ──────────────────────────────────────────────────────────────────────────────

def create_app() -> Flask:
    """Fabrique l'application Flask (factory pattern, testable)."""
    app = Flask(__name__)
    # Évite les surprises UTF-8 dans les réponses JSON (accents PMSI)
    app.config["JSON_AS_ASCII"] = False

    # ──────────────────────────────────────────────────────────────────────
    # CORS — En-têtes ajoutés à chaque réponse
    # ──────────────────────────────────────────────────────────────────────
    @app.after_request
    def _cors(response):
        origin = request.headers.get("Origin", "")
        if ALLOWED_ORIGINS == ["*"]:
            response.headers["Access-Control-Allow-Origin"] = "*"
        elif origin in ALLOWED_ORIGINS:
            response.headers["Access-Control-Allow-Origin"] = origin
            response.headers["Vary"] = "Origin"
        response.headers["Access-Control-Allow-Methods"] = "GET, POST, OPTIONS"
        response.headers["Access-Control-Allow-Headers"] = (
            "Content-Type, Authorization"
        )
        return response

    # Préflight CORS pour les POST cross-origin depuis le PHP
    @app.route("/<path:_any>", methods=["OPTIONS"])
    def _preflight(_any):  # noqa: ARG001 — wildcard
        return ("", 204)

    # ──────────────────────────────────────────────────────────────────────
    # AUTH — Jeton Bearer partagé
    # ──────────────────────────────────────────────────────────────────────
    def _require_token(view: Callable) -> Callable:
        """Décorateur : rejette toute requête sans le bon jeton Bearer."""

        @wraps(view)
        def wrapper(*args, **kwargs):
            # Auth désactivée si aucun jeton n'est défini (dev local)
            if not BRIDGE_TOKEN:
                return view(*args, **kwargs)
            header = request.headers.get("Authorization", "")
            if not header.startswith("Bearer "):
                return jsonify(error="Missing bearer token"), 401
            token = header[len("Bearer ") :].strip()
            # Comparaison constante pour éviter les timing attacks
            if not secrets.compare_digest(token, BRIDGE_TOKEN):
                return jsonify(error="Invalid token"), 403
            return view(*args, **kwargs)

        return wrapper

    # ──────────────────────────────────────────────────────────────────────
    # HEALTHCHECK — public, utilisé par le PHP pour vérifier la liaison
    # ──────────────────────────────────────────────────────────────────────
    @app.get("/health")
    def health():
        return jsonify(
            status="ok",
            service="sovereign-os-dim",
            version="V32.0",
            auth_required=bool(BRIDGE_TOKEN),
            formats=len(_processor.matrix),
        )

    # ──────────────────────────────────────────────────────────────────────
    # MATRIX & IDENTIFICATION — Lecture seule (stateless)
    # ──────────────────────────────────────────────────────────────────────
    @app.get("/api/matrix")
    @_require_token
    def api_matrix():
        return jsonify(
            [
                {
                    "key": k,
                    "length": v["length"],
                    "ipp": list(v["ipp"]),
                    "ddn": list(v["ddn"]),
                    "desc": v.get("desc", ""),
                    "field": v.get("field", ""),
                    "since": v.get("since"),
                }
                for k, v in _processor.matrix.items()
            ]
        )

    @app.post("/api/identify")
    @_require_token
    def api_identify():
        payload = request.get_json(silent=True) or {}
        name = payload.get("filename") or payload.get("path") or ""
        if not name:
            return jsonify(error="filename requis"), 400
        fmt = DataProcessor.identify_format(name)
        spec = _processor.matrix.get(fmt) if fmt else None
        return jsonify(
            filename=os.path.basename(name),
            format=fmt,
            spec=spec
            and {
                "length": spec["length"],
                "ipp": list(spec["ipp"]),
                "ddn": list(spec["ddn"]),
                "field": spec.get("field"),
            },
        )

    # ──────────────────────────────────────────────────────────────────────
    # SCAN & PROCESS
    # ──────────────────────────────────────────────────────────────────────
    @app.post("/api/scan")
    @_require_token
    def api_scan():
        global _current_files
        payload = request.get_json(silent=True) or {}
        folders = payload.get("folders") or []
        if not isinstance(folders, list) or not folders:
            return jsonify(error="folders (list) requis"), 400

        valid = [f for f in folders if isinstance(f, str) and os.path.isdir(f)]
        if not valid:
            return jsonify(error="Aucun dossier valide", folders=folders), 400

        _current_folders.clear()
        _current_folders.extend(valid)
        _current_files = _processor.scan_multiple_directories(valid)
        return jsonify(
            files=_current_files,
            count=len(_current_files),
            logs=_processor.get_logs(),
        )

    @app.post("/api/process")
    @_require_token
    def api_process():
        """Scan + process en un appel (équivalent scan_and_process)."""
        global _current_files
        payload = request.get_json(silent=True) or {}
        folders = payload.get("folders") or _current_folders

        if folders:
            valid = [
                f for f in folders if isinstance(f, str) and os.path.isdir(f)
            ]
            if not valid:
                return jsonify(error="Aucun dossier valide"), 400
            _current_folders.clear()
            _current_folders.extend(valid)
            _current_files = _processor.scan_multiple_directories(valid)

        if not _current_files:
            return jsonify(error="Aucun fichier à traiter"), 400

        stats = _processor.process_files(_current_files)
        return jsonify(
            files=_current_files,
            stats=stats,
            logs=_processor.get_logs(),
        )

    # ──────────────────────────────────────────────────────────────────────
    # MPI — Collisions & résolution
    # ──────────────────────────────────────────────────────────────────────
    @app.get("/api/collisions")
    @_require_token
    def api_collisions():
        query = (request.args.get("q") or "").strip().lower()
        cols = _processor.get_collisions()
        if query:
            cols = [c for c in cols if query in c["ipp"].lower()]
        return jsonify(collisions=cols, count=len(cols))

    @app.post("/api/resolve")
    @_require_token
    def api_resolve():
        """
        Résolution d'une collision.
          - {"auto": true}                 → auto_resolve_all
          - {"ipp": "...", "ddn": "..."}   → set_pivot manuel
        """
        payload = request.get_json(silent=True) or {}

        if payload.get("auto"):
            count = _processor.auto_resolve_all()
            return jsonify(
                resolved=count,
                collisions=_processor.get_collisions(),
                mpi_stats=_processor.get_mpi_stats(),
                logs=_processor.get_logs(),
            )

        ipp = payload.get("ipp")
        ddn = payload.get("ddn")
        if not ipp or not ddn:
            return jsonify(error="ipp et ddn requis (ou auto=true)"), 400

        ok = _processor.set_pivot(ipp, ddn)
        return jsonify(
            success=ok,
            mpi_stats=_processor.get_mpi_stats(),
            logs=_processor.get_logs(),
        )

    # ──────────────────────────────────────────────────────────────────────
    # STATS
    # ──────────────────────────────────────────────────────────────────────
    @app.get("/api/stats")
    @_require_token
    def api_stats():
        mpi = _processor.get_mpi_stats()
        return jsonify(
            folders=len(_current_folders),
            files=len(_current_files),
            formats=len(_processor.matrix),
            mpi=mpi,
            format_breakdown=_processor.get_format_breakdown(),
            file_stats=_processor.file_stats,
        )

    # ──────────────────────────────────────────────────────────────────────
    # EXPORT
    # ──────────────────────────────────────────────────────────────────────
    @app.post("/api/export")
    @_require_token
    def api_export():
        payload = request.get_json(silent=True) or {}
        output_dir = payload.get("output_dir")
        if not _current_files:
            return jsonify(error="Aucun fichier"), 400
        if not output_dir:
            if not _current_folders:
                return jsonify(error="output_dir requis"), 400
            output_dir = os.path.join(_current_folders[0], "PILOT_OUTPUT")

        result = _processor.export_csv(_current_files, output_dir)
        result["logs"] = _processor.get_logs()
        return jsonify(result)

    @app.post("/api/export-sanitized")
    @_require_token
    def api_export_sanitized():
        payload = request.get_json(silent=True) or {}
        filepath = payload.get("path")
        output_dir = payload.get("output_dir")
        if not filepath:
            return jsonify(error="path requis"), 400
        if not output_dir:
            if not _current_folders:
                return jsonify(error="output_dir requis"), 400
            output_dir = os.path.join(_current_folders[0], "PILOT_OUTPUT")
        result = _processor.export_sanitized_txt(filepath, output_dir)
        return jsonify(result)

    # ──────────────────────────────────────────────────────────────────────
    # INSPECTION & IMPORT CSV
    # ──────────────────────────────────────────────────────────────────────
    @app.post("/api/inspect")
    @_require_token
    def api_inspect():
        payload = request.get_json(silent=True) or {}
        filepath = payload.get("path")
        if not filepath:
            return jsonify(error="path requis"), 400
        return jsonify(_processor.inspect_file(filepath))

    @app.post("/api/import-csv")
    @_require_token
    def api_import_csv():
        import csv

        payload = request.get_json(silent=True) or {}
        filepath = payload.get("path")
        if not filepath or not os.path.isfile(filepath):
            return jsonify(error="Fichier introuvable"), 404
        try:
            with open(filepath, "r", encoding="utf-8", errors="replace") as f:
                sample = f.read(4096)
                try:
                    dialect = csv.Sniffer().sniff(sample, delimiters=";,\t")
                    sep = dialect.delimiter
                except csv.Error:
                    sep = ";"
            rows, headers = [], []
            with open(filepath, "r", encoding="utf-8", errors="replace") as f:
                reader = csv.reader(f, delimiter=sep)
                for i, row in enumerate(reader):
                    if i == 0:
                        headers = row
                        continue
                    if i > 5000:
                        break
                    rows.append(row)
            return jsonify(
                filename=os.path.basename(filepath),
                path=filepath,
                headers=headers,
                rows=rows,
                total_rows=len(rows),
                separator=sep,
            )
        except (OSError, UnicodeDecodeError) as e:
            return jsonify(error=str(e)), 500

    # ──────────────────────────────────────────────────────────────────────
    # IMPORT EXCEL + VISUALISATION
    # ──────────────────────────────────────────────────────────────────────
    # Pourquoi ? Les équipes DIM reçoivent régulièrement des tableaux Excel
    # (suivi mensuel, extractions BigQuery, retours ATIH). Le bridge lit ces
    # classeurs et renvoie soit la table brute, soit une série agrégée prête
    # à être tracée par Chart.js côté PHP (option `chart-from-excel`).

    def _read_excel(path: str, sheet: str | None = None, limit: int = 5000):
        """
        Charge un classeur Excel et retourne (headers, rows, sheets).

        On lit en `values_only` pour éviter les références de cellules et on
        coerce en str/float/int suivant le type natif — les dates sont
        retournées en ISO pour être JSON-serializable.
        """
        try:
            from openpyxl import load_workbook
        except ImportError as e:
            raise RuntimeError(
                "openpyxl n'est pas installé — pip install openpyxl>=3.1"
            ) from e

        from datetime import date, datetime as _dt

        wb = load_workbook(path, read_only=True, data_only=True)
        sheets = wb.sheetnames
        ws = wb[sheet] if sheet and sheet in sheets else wb[sheets[0]]

        headers: list[str] = []
        rows: list[list] = []
        for i, row in enumerate(ws.iter_rows(values_only=True)):
            if i == 0:
                headers = [str(c) if c is not None else "" for c in row]
                continue
            if i > limit:
                break
            cleaned = []
            for c in row:
                if isinstance(c, (_dt, date)):
                    cleaned.append(c.isoformat())
                elif c is None:
                    cleaned.append("")
                else:
                    cleaned.append(c)
            rows.append(cleaned)
        wb.close()
        return headers, rows, sheets

    @app.post("/api/import-excel")
    @_require_token
    def api_import_excel():
        payload = request.get_json(silent=True) or {}
        filepath = payload.get("path")
        sheet = payload.get("sheet")
        limit = int(payload.get("limit") or 5000)
        if not filepath or not os.path.isfile(filepath):
            return jsonify(error="Fichier introuvable"), 404
        try:
            headers, rows, sheets = _read_excel(filepath, sheet, limit)
        except RuntimeError as e:
            return jsonify(error=str(e)), 500
        except Exception as e:  # openpyxl peut lever InvalidFileException
            return jsonify(error=f"Lecture impossible : {e}"), 400

        return jsonify(
            filename=os.path.basename(filepath),
            path=filepath,
            sheet=sheet or (sheets[0] if sheets else None),
            sheets=sheets,
            headers=headers,
            rows=rows,
            total_rows=len(rows),
        )

    @app.post("/api/chart-from-excel")
    @_require_token
    def api_chart_from_excel():
        """
        Construit une (ou plusieurs) séries agrégées à partir d'un ou de
        plusieurs classeurs Excel, prêtes pour Chart.js.

        Payload (entrée) — deux modes :

          Mode 1 (un seul fichier, rétro-compatible) :
            {
              "path":  "/chemin/fichier.xlsx",
              "sheet": "Feuil1",                 (optionnel)
              "label": "Mois",                   (colonne X — obligatoire)
              "value": "Séjours",                (colonne numérique — obligatoire)
              "agg":   "sum" | "avg" | "count",  (défaut: sum)
              "top":   20                         (N buckets max, 0 = tous)
            }

          Mode 2 (plusieurs fichiers, comparaison ou fusion) :
            {
              "paths":  ["/a.xlsx", "/b.xlsx", ...],
              "sheet":  null | "NomFeuille",
              "label":  "Mois",
              "value":  "Séjours",
              "agg":    "sum",
              "top":    20,
              "mode":   "merge" | "compare"      (défaut: compare)
            }
            - merge   → une seule série agrégée sur TOUS les fichiers
            - compare → une série par fichier, alignée sur l'union des labels

        Réponse :
            {
              "agg": "...", "label_col": "...", "value_col": "...",
              "labels": [...],
              "datasets": [
                 { "name": "fichier.xlsx", "values": [...] }, ...
              ],
              # Compat Chart.js simple : si une seule série, on renvoie aussi "values"
              "values": [...]
            }
        """
        payload = request.get_json(silent=True) or {}

        # Normalise en liste de chemins : accepte "path" (str) ou "paths" (list)
        raw_paths = payload.get("paths")
        if not raw_paths:
            single = payload.get("path")
            raw_paths = [single] if single else []
        if isinstance(raw_paths, str):
            raw_paths = [raw_paths]

        paths = [
            p for p in raw_paths if isinstance(p, str) and os.path.isfile(p)
        ]
        if not paths:
            return jsonify(error="Aucun fichier Excel valide trouvé"), 404

        label_col = payload.get("label")
        value_col = payload.get("value")
        if not label_col or not value_col:
            return jsonify(error="label et value requis"), 400

        agg = (payload.get("agg") or "sum").lower()
        if agg not in ("sum", "avg", "count"):
            return jsonify(error="agg doit être sum, avg ou count"), 400
        top = int(payload.get("top") or 20)
        sheet = payload.get("sheet")
        mode = (payload.get("mode") or ("merge" if len(paths) == 1 else "compare")).lower()
        if mode not in ("merge", "compare"):
            return jsonify(error="mode doit être merge ou compare"), 400

        # Lecture de chaque fichier et construction de buckets par fichier
        # per_file[filename] = { label: [values...] }
        per_file: dict[str, dict[str, list[float]]] = {}
        effective_agg = agg  # peut passer en "count" si la colonne n'est pas num

        for fp in paths:
            try:
                headers, rows, _sheets = _read_excel(fp, sheet, limit=50000)
            except RuntimeError as e:
                return jsonify(error=str(e)), 500
            except Exception as e:
                return jsonify(error=f"{os.path.basename(fp)} : {e}"), 400

            if label_col not in headers or value_col not in headers:
                return jsonify(
                    error=f"Colonnes introuvables dans {os.path.basename(fp)}",
                    available=headers,
                ), 400

            li = headers.index(label_col)
            vi = headers.index(value_col)

            file_buckets: dict[str, list[float]] = {}
            for row in rows:
                if li >= len(row) or vi >= len(row):
                    continue
                key = "" if row[li] is None else str(row[li])
                raw = row[vi]
                try:
                    num = float(raw) if raw not in (None, "") else 0.0
                except (TypeError, ValueError):
                    num = 1.0
                    effective_agg = "count"
                file_buckets.setdefault(key, []).append(num)

            per_file[os.path.basename(fp)] = file_buckets

        def _reduce(values: list[float]) -> float:
            if not values:
                return 0.0
            if effective_agg == "sum":
                return sum(values)
            if effective_agg == "avg":
                return sum(values) / len(values)
            return float(len(values))  # count

        # ─── Mode MERGE : on fusionne toutes les listes par label ─────────
        if mode == "merge" or len(paths) == 1:
            merged: dict[str, list[float]] = {}
            for buckets in per_file.values():
                for label, vals in buckets.items():
                    merged.setdefault(label, []).extend(vals)

            series = [(k, _reduce(v)) for k, v in merged.items()]
            series.sort(key=lambda x: x[1], reverse=True)
            if top > 0:
                series = series[:top]

            return jsonify(
                mode="merge",
                agg=effective_agg,
                label_col=label_col,
                value_col=value_col,
                files=list(per_file.keys()),
                labels=[s[0] for s in series],
                values=[s[1] for s in series],
                datasets=[
                    {
                        "name": "Total",
                        "values": [s[1] for s in series],
                    }
                ],
                total_buckets=len(merged),
            )

        # ─── Mode COMPARE : une série par fichier, alignée sur les labels ─
        # Union des labels, puis tri par somme globale décroissante
        label_totals: dict[str, float] = {}
        for buckets in per_file.values():
            for label, vals in buckets.items():
                label_totals[label] = label_totals.get(label, 0.0) + _reduce(vals)

        ordered_labels = sorted(label_totals, key=lambda k: label_totals[k], reverse=True)
        if top > 0:
            ordered_labels = ordered_labels[:top]

        datasets = []
        for filename, buckets in per_file.items():
            datasets.append(
                {
                    "name": filename,
                    "values": [
                        _reduce(buckets.get(lbl, [])) for lbl in ordered_labels
                    ],
                }
            )

        return jsonify(
            mode="compare",
            agg=effective_agg,
            label_col=label_col,
            value_col=value_col,
            files=list(per_file.keys()),
            labels=ordered_labels,
            # Pour compat : "values" = première série (utile pour pie/doughnut)
            values=datasets[0]["values"] if datasets else [],
            datasets=datasets,
            total_buckets=len(label_totals),
        )

    # ──────────────────────────────────────────────────────────────────────
    # RESET
    # ──────────────────────────────────────────────────────────────────────
    @app.post("/api/reset")
    @_require_token
    def api_reset():
        _reset_state()
        return jsonify(ok=True)

    # ──────────────────────────────────────────────────────────────────────
    # ERROR HANDLERS — JSON plutôt que HTML
    # ──────────────────────────────────────────────────────────────────────
    @app.errorhandler(404)
    def _nf(_e):
        return jsonify(error="Not found"), 404

    @app.errorhandler(405)
    def _na(_e):
        return jsonify(error="Method not allowed"), 405

    @app.errorhandler(500)
    def _ie(e):
        return jsonify(error="Internal server error", detail=str(e)), 500

    return app


# ──────────────────────────────────────────────────────────────────────────────
# CLI
# ──────────────────────────────────────────────────────────────────────────────

def main():
    parser = argparse.ArgumentParser(
        description="Sovereign OS DIM — HTTP REST bridge (PHP integration)"
    )
    parser.add_argument("--host", default=DEFAULT_HOST)
    parser.add_argument("--port", type=int, default=DEFAULT_PORT)
    parser.add_argument(
        "--debug",
        action="store_true",
        help="Active le reloader Flask (dev only)",
    )
    args = parser.parse_args()

    app = create_app()
    banner = (
        "╔══════════════════════════════════════════════════════════╗\n"
        "║   🛡️  SOVEREIGN OS DIM — HTTP BRIDGE                    ║\n"
        "╚══════════════════════════════════════════════════════════╝\n"
        f"  Listening on http://{args.host}:{args.port}\n"
        f"  Auth: {'Bearer token required' if BRIDGE_TOKEN else 'DISABLED (dev)'}\n"
        f"  CORS origins: {', '.join(ALLOWED_ORIGINS) or '(none)'}\n"
    )
    print(banner)
    app.run(host=args.host, port=args.port, debug=args.debug)


if __name__ == "__main__":
    main()
