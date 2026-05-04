"""
═══════════════════════════════════════════════════════════════════════════════
 backend/ml/parse_atih_specs.py · Parser des specs Excel officielles ATIH 2026
═══════════════════════════════════════════════════════════════════════════════

Lit les classeurs Excel publiés par l'ATIH (formats_psy_2026, formats_mco_2026,
formats_smr_2026, formats_had_2026, et leurs variantes anonymes/visual) et
extrait pour chaque format ·

  - Nom du recueil (RPS, RAA, FICHSUP-PSY, ...)
  - Secteur (DAF / ex-OQN / les deux)
  - Longueur de ligne (formule "154 + 8*nDA + 23*nZA" → base 154)
  - Position et taille de chaque champ (IPP, DDN, FINESS, ...)

Ces specs réelles remplacent les estimations dans synthetic.py.

Usage ·
    python -m backend.ml.parse_atih_specs --src C:/Users/adamb/Downloads \\
        --out backend/ml/data/atih_specs_2026.json
═══════════════════════════════════════════════════════════════════════════════
"""

from __future__ import annotations

import argparse
import glob
import json
import os
import re
from dataclasses import asdict, dataclass, field
from typing import Any

import openpyxl


# ─────────────────────────────────────────────────────────────────────────────
# MODÈLES
# ─────────────────────────────────────────────────────────────────────────────


@dataclass
class FieldSpec:
    label: str
    size: int
    start: int  # 1-indexed comme dans les specs ATIH
    end: int
    type_: str = ""
    required: str = ""


@dataclass
class FormatSpec2026:
    name: str
    field: str  # PSY / MCO / SMR / HAD
    sector: str = ""  # DAF / ex-OQN / both / ""
    base_length: int = 0
    extension_formula: str = ""  # ex "+ 8*nDA + 23*nZA"
    fields: list[FieldSpec] = field(default_factory=list)
    source_file: str = ""
    source_sheet: str = ""

    @property
    def ipp(self) -> FieldSpec | None:
        return self._find_field(["IPP", "Identifiant Permanent du Patient",
                                 "N� d'identification permanent du patient"])

    @property
    def ddn(self) -> FieldSpec | None:
        return self._find_field(["Date de naissance", "DDN"])

    @property
    def finess_geo(self) -> FieldSpec | None:
        return self._find_field(["FINESS g�ographique", "FINESS geographique",
                                 "N� FINESS g�ographique"])

    def _find_field(self, candidates: list[str]) -> FieldSpec | None:
        for f in self.fields:
            label_low = f.label.lower() if f.label else ""
            for c in candidates:
                if c.lower() in label_low:
                    return f
        return None


# ─────────────────────────────────────────────────────────────────────────────
# PARSER
# ─────────────────────────────────────────────────────────────────────────────


_FIELD_NAME_FROM_FILENAME = {
    "psy": "PSY",
    "mco": "MCO",
    "smr": "SMR",
    "had": "HAD",
}


def _detect_field_from_filename(path: str) -> str:
    base = os.path.basename(path).lower()
    for key, val in _FIELD_NAME_FROM_FILENAME.items():
        if key in base:
            return val
    return "unknown"


def _parse_length_line(text: str) -> tuple[int, str]:
    """
    Extrait la longueur de base et la formule d'extension depuis ·
        'Nombre de caract�res attendus pour un enregistrement = 154 + (8*nDA) + (23*nZA)'
    Retourne (154, '+ (8*nDA) + (23*nZA)').
    """
    if not text:
        return 0, ""
    m = re.search(r"=\s*(\d+)\s*(.*)", text)
    if not m:
        return 0, ""
    base = int(m.group(1))
    rest = m.group(2).strip()
    return base, rest


def _is_field_header_row(row: tuple) -> bool:
    """Détecte la ligne d'en-tête (Libell�, Taille, D�but, Fin, ...)."""
    if not row:
        return False
    cells = [str(c) if c is not None else "" for c in row]
    joined = " ".join(cells).lower()
    return ("libell" in joined or "libelle" in joined) and "taille" in joined and \
           ("d�but" in joined or "debut" in joined or "deb" in joined)


def _parse_field_row(row: tuple, header_offset: int) -> FieldSpec | None:
    """
    Parse une ligne de champ. Le header peut commencer par 'Libell�' (col 0)
    ou avoir une colonne fantôme avant. On laisse parler les valeurs · si
    c'est un int, c'est probablement Taille / Début / Fin.
    """
    cells = list(row)
    # Trouver les 3 premiers entiers consécutifs · taille, debut, fin
    label = None
    nums = []
    for i, c in enumerate(cells):
        if isinstance(c, int) and 1 <= c <= 99999:
            nums.append((i, c))
        elif isinstance(c, str) and c.strip() and label is None:
            label = c.strip()
    if len(nums) < 3 or not label:
        return None
    # taille / début / fin · les 3 premiers entiers
    size, start, end = nums[0][1], nums[1][1], nums[2][1]
    # Type · 1 cellule string courte après 'fin'
    type_ = ""
    type_idx = nums[2][0] + 1
    if type_idx < len(cells) and isinstance(cells[type_idx], str) \
            and len(cells[type_idx].strip()) <= 4:
        type_ = cells[type_idx].strip()
    # Required · cellule contenant 'O' / 'F' / 'C'
    required = ""
    for c in cells[type_idx + 1:]:
        if isinstance(c, str) and c.strip() in ("O", "F", "C", "Oui", "Non"):
            required = c.strip()
            break
    return FieldSpec(label=label, size=size, start=start, end=end,
                     type_=type_, required=required)


def _parse_sheet(ws, file_field: str, source_file: str) -> FormatSpec2026 | None:
    sheet_name = ws.title.strip()
    # Skip sheets that are clearly not format specs
    if any(skip in sheet_name.lower() for skip in
           ["pr�sentation", "presentation", "modifications", "lisez"]):
        return None
    # Lit les 5 premières lignes pour trouver le titre + sector + length
    title = ""
    sector = ""
    base_length = 0
    extension = ""
    rows = list(ws.iter_rows(min_row=1, max_row=15, values_only=True))
    for row in rows:
        for cell in row:
            if not isinstance(cell, str):
                continue
            txt = cell.strip()
            if not title and ("Format " in txt or "FICHIER" in txt.upper()):
                title = txt
            if not sector and txt.upper() in ("DAF", "EX-OQN", "OQN", "DAF/EX-OQN",
                                              "DAF/EX-OQN ", "DAF / EX-OQN",
                                              "DAF/OQN"):
                sector = txt
            if base_length == 0 and "Nombre de caract" in txt and "=" in txt:
                base_length, extension = _parse_length_line(txt)
    name = title.replace("Format ", "").strip() if title else sheet_name.strip()

    # Identifier la ligne d'en-tête de tableau
    header_row_idx = None
    for i, row in enumerate(rows):
        if _is_field_header_row(row):
            header_row_idx = i
            break
    if header_row_idx is None:
        # Pas de tableau détecté · on retourne quand même la spec basique
        return FormatSpec2026(
            name=name, field=file_field, sector=sector,
            base_length=base_length, extension_formula=extension,
            source_file=source_file, source_sheet=sheet_name,
        )

    # Parser tous les champs après le header
    fields_list: list[FieldSpec] = []
    for row in ws.iter_rows(min_row=header_row_idx + 2, values_only=True):
        if all(c is None for c in row):
            continue
        fs = _parse_field_row(row, header_offset=header_row_idx)
        if fs:
            fields_list.append(fs)

    return FormatSpec2026(
        name=name, field=file_field, sector=sector,
        base_length=base_length, extension_formula=extension,
        fields=fields_list,
        source_file=source_file, source_sheet=sheet_name,
    )


def parse_workbook(path: str) -> list[FormatSpec2026]:
    file_field = _detect_field_from_filename(path)
    wb = openpyxl.load_workbook(path, data_only=True)
    out: list[FormatSpec2026] = []
    for sheet in wb.sheetnames:
        try:
            spec = _parse_sheet(wb[sheet], file_field, source_file=os.path.basename(path))
            if spec and (spec.base_length > 0 or spec.fields):
                out.append(spec)
        except Exception as e:  # pragma: no cover
            print(f"[WARN] {os.path.basename(path)}::{sheet} · {e}")
    return out


# ─────────────────────────────────────────────────────────────────────────────
# CLI
# ─────────────────────────────────────────────────────────────────────────────


def main() -> None:  # pragma: no cover
    p = argparse.ArgumentParser(description="Parse les Excel ATIH 2026.")
    p.add_argument("--src", default="C:/Users/adamb/Downloads",
                   help="Dossier contenant les fichiers formats_*_2026*.xlsx")
    p.add_argument("--out", default="backend/ml/data/atih_specs_2026.json")
    args = p.parse_args()

    files = sorted(glob.glob(os.path.join(args.src, "formats_*_2026*.xlsx")))
    if not files:
        print(f"[ERR] Aucun fichier formats_*_2026*.xlsx dans {args.src}")
        raise SystemExit(1)

    print(f"[ATIH] {len(files)} fichiers a traiter ·")
    all_specs: list[dict[str, Any]] = []
    for f in files:
        specs = parse_workbook(f)
        psy_count = sum(1 for s in specs if s.field == "PSY")
        print(f"  {os.path.basename(f):<40} · {len(specs)} formats "
              f"({psy_count} PSY)")
        for s in specs:
            d = asdict(s)
            d["fields"] = [asdict(fld) for fld in s.fields]
            d["ipp_position"] = (s.ipp.start, s.ipp.size) if s.ipp else None
            d["ddn_position"] = (s.ddn.start, s.ddn.size) if s.ddn else None
            d["finess_position"] = (s.finess_geo.start, s.finess_geo.size) \
                                    if s.finess_geo else None
            all_specs.append(d)

    os.makedirs(os.path.dirname(args.out), exist_ok=True)
    with open(args.out, "w", encoding="utf-8") as f:
        json.dump(all_specs, f, ensure_ascii=False, indent=2)

    n_with_length = sum(1 for s in all_specs if s["base_length"] > 0)
    n_with_ipp = sum(1 for s in all_specs if s["ipp_position"])
    n_with_ddn = sum(1 for s in all_specs if s["ddn_position"])
    print(f"\n[ATIH] {len(all_specs)} formats parses au total")
    print(f"        {n_with_length} avec longueur de base detectee")
    print(f"        {n_with_ipp} avec position IPP detectee")
    print(f"        {n_with_ddn} avec position DDN detectee")
    print(f"        sortie · {args.out}")


if __name__ == "__main__":  # pragma: no cover
    main()
